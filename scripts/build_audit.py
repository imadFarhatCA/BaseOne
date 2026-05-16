#!/usr/bin/env python3
"""Build a content-audit Excel file for the Base One website.

Reads every page/component and the relevant data files, extracts user-visible
text, assigns stable IDs, and writes:
  - content-audit.xlsx  (one sheet per page; auditor edits "New Text" column)
  - content-audit-manifest.json  (id -> file path + exact original text)

After the auditor returns the Excel, a sister script will use the manifest to
replace each modified text in the source files.
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path("/Users/imadfarhat/imadFarhat projects/Base One")
SRC = ROOT / "src"
OUT_XLSX = ROOT / "content-audit.xlsx"
OUT_MANIFEST = ROOT / "content-audit-manifest.json"

# ────────────────────────────────────────────────────────────────────────────
# Page configuration — order matters (becomes sheet order)
# Each page has its own ID prefix.
# ────────────────────────────────────────────────────────────────────────────
PAGES = [
    {"key": "HOME",  "name": "Home",            "files": ["routes/+page.svelte"]},
    {"key": "DIVE",  "name": "Diving",          "files": ["routes/diving/+page.svelte", "lib/data/caves.js"]},
    {"key": "TRAIN", "name": "Training",        "files": ["routes/training/+page.svelte"]},
    {"key": "EXPL",  "name": "Exploration",     "files": ["routes/exploration/+page.svelte"]},
    {"key": "FAC",   "name": "The Facility",    "files": ["routes/about/+page.svelte", "lib/data/team.js"]},
    {"key": "CALA",  "name": "Cala Gonone",     "files": ["routes/cala-gonone/+page.svelte", "lib/data/area.js"]},
    {"key": "PLAN",  "name": "Plan Your Trip",  "files": ["routes/plan/+page.svelte", "lib/data/simulator.js"]},
    {"key": "SHARED","name": "Shared (Nav · Footer · CTA · Hero)", "files": [
        "lib/components/Nav.svelte",
        "lib/components/Footer.svelte",
        "lib/components/CtaBlock.svelte",
        "lib/components/PageHero.svelte",
    ]},
]

# ────────────────────────────────────────────────────────────────────────────
# Section detection
# ────────────────────────────────────────────────────────────────────────────
SECTION_COMMENT_RE = re.compile(r"<!--\s*[─\-]+\s*(.*?)\s*[─\-]+\s*-->")
SECTION_PLAIN_RE   = re.compile(r"<!--\s*(.*?)\s*-->")

# ────────────────────────────────────────────────────────────────────────────
# Text-bearing HTML tags. Order matters slightly for nested matches.
# ────────────────────────────────────────────────────────────────────────────
TEXT_TAGS = ["h1", "h2", "h3", "h4", "h5", "p", "li", "button", "a",
             "span", "blockquote", "summary", "label"]

# Inline elements that may wrap text inside a paragraph — handled by capturing
# the parent text and stripping inner tags.
INLINE_TAGS_TO_FLATTEN = {"strong", "em", "b", "i", "u", "small", "br"}

# Attributes worth auditing (alt text, aria, placeholder, title).
TEXT_ATTRS = ["alt", "title", "placeholder", "aria-label"]

# Component props that carry text (PageHero, CtaBlock, Testimonials etc.)
COMPONENT_TEXT_PROPS = ["heading", "eyebrow", "sub", "text", "primaryLabel", "secondaryLabel"]

# Skip-scripts/styles inside .svelte
SCRIPT_RE = re.compile(r"<script[^>]*>.*?</script>", re.S)
STYLE_RE  = re.compile(r"<style[^>]*>.*?</style>", re.S)
COMMENT_RE = re.compile(r"<!--.*?-->", re.S)


def strip_inline_tags(html: str) -> str:
    """Remove inline tags like <strong>, <em>, <br>, keeping the text."""
    out = re.sub(r"<br\s*/?>", " ", html)
    out = re.sub(r"</?\s*(strong|em|b|i|u|small)\s*[^>]*>", "", out, flags=re.I)
    return out


def clean_text(s: str) -> str:
    if s is None: return ""
    # Decode common entities
    s = s.replace("&amp;", "&").replace("&nbsp;", " ").replace("&mdash;", "—") \
         .replace("&ndash;", "–").replace("&apos;", "'").replace("&quot;", '"') \
         .replace("&#39;", "'")
    # Strip Svelte directives like {#if ...}, {/if}, {#each ...}, {:else}
    s = re.sub(r"\{[#:/][^}]*\}", " ", s)
    # Strip Svelte expressions like {variable} unless they're standalone
    s = re.sub(r"\{[^}]*\}", " ", s)
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


def looks_like_text(s: str) -> bool:
    """Decide whether a candidate string is real user-visible text."""
    if not s: return False
    if len(s) < 2: return False
    if s.isdigit(): return False
    # Pure URL or path
    if re.match(r"^(https?://|/|\.\.?/|#|mailto:|tel:|wa\.me|javascript:)", s): return False
    if re.match(r"^[a-z][a-z0-9-]*\.(svg|png|jpg|jpeg|webp|gif|css|js|json)$", s, re.I): return False
    # CSS class lists
    if re.match(r"^[a-z][a-z0-9_-]*(\s+[a-z][a-z0-9_-]*)+$", s) and not " " in s.strip(" "): return False
    # Only punctuation / symbols
    if re.match(r"^[\W_]+$", s): return False
    return True


def extract_svelte(path: Path):
    """Yield (section, element_type, text) tuples from a .svelte file."""
    raw = path.read_text(encoding="utf-8")
    # Remove script and style blocks
    body = SCRIPT_RE.sub("", raw)
    body = STYLE_RE.sub("", body)

    # Walk line by line; track current section from comments
    lines = body.split("\n")
    section = "(top of page)"
    buf = []
    for ln in lines:
        m = SECTION_COMMENT_RE.search(ln)
        if m:
            section = m.group(1).strip(" ─-:")
            if not section: section = "(top of page)"
            continue
        # Also accept "<!-- Section -->" plain
        if not m and "<!--" in ln and "-->" in ln:
            m2 = SECTION_PLAIN_RE.search(ln)
            if m2:
                cand = m2.group(1).strip()
                # Treat short, title-like comments as a section marker
                if 3 <= len(cand) <= 60 and cand[0].isupper():
                    section = cand
                    continue
        buf.append((section, ln))

    # Now scan for text in heading/paragraph/list/button/link/span and component props.
    for section, ln in buf:
        # Component props (multi-attribute, may span lines but we look per line)
        for prop in COMPONENT_TEXT_PROPS:
            for m in re.finditer(rf'{prop}\s*=\s*"([^"]+)"', ln):
                t = clean_text(strip_inline_tags(m.group(1)))
                if looks_like_text(t):
                    yield section, f"prop:{prop}", t, m.group(1)

        # alt/title/placeholder/aria-label
        for attr in TEXT_ATTRS:
            for m in re.finditer(rf'\b{attr}\s*=\s*"([^"]+)"', ln):
                t = clean_text(strip_inline_tags(m.group(1)))
                if looks_like_text(t):
                    yield section, f"attr:{attr}", t, m.group(1)

        # Text inside heading/paragraph/list/button/link/span/blockquote/summary
        for tag in TEXT_TAGS:
            for m in re.finditer(
                rf"<{tag}\b[^>]*>(.*?)</{tag}>", ln, flags=re.I
            ):
                inner = strip_inline_tags(m.group(1))
                # Skip if the inner contains another block tag (avoid double extraction)
                if re.search(r"<(h[1-6]|p|li|button|blockquote|section|div|ul|ol)\b", inner, re.I):
                    continue
                t = clean_text(inner)
                if looks_like_text(t):
                    yield section, f"<{tag}>", t, m.group(1)


def extract_js_data(path: Path):
    """Extract user-visible strings from data files (team.js, area.js, etc.)."""
    raw = path.read_text(encoding="utf-8")
    # Match field: 'value' or field: "value" — focusing on text-bearing fields
    text_fields = {
        "name", "role", "title", "desc", "label", "sub", "text", "heading",
        "time", "notes", "from", "to", "via", "eyebrow",
    }
    # Single-line strings
    pattern = re.compile(
        r"\b(" + "|".join(text_fields) + r")\s*:\s*(['\"])((?:\\\2|(?!\2).)*)\2",
        re.S,
    )
    # Track context — find the nearest enclosing top-level "export const X = ["
    export_blocks = []
    for m in re.finditer(r"export const (\w+)\s*=", raw):
        export_blocks.append((m.start(), m.group(1)))

    def section_for(pos: int) -> str:
        last = "(top)"
        for start, name in export_blocks:
            if start <= pos:
                last = name
        return last

    for m in pattern.finditer(raw):
        field = m.group(1)
        val = clean_text(strip_inline_tags(m.group(3)))
        if looks_like_text(val):
            yield section_for(m.start()), f"data:{field}", val, m.group(3)

    # Multi-line bio arrays inside team.js: bio: [ '...', '...' ]
    for m in re.finditer(r"\bbio\s*:\s*\[(.*?)\]", raw, re.S):
        for s in re.finditer(r"(['\"])((?:\\\1|(?!\1).)*)\1", m.group(1)):
            val = clean_text(strip_inline_tags(s.group(2)))
            if looks_like_text(val):
                yield section_for(m.start()), "data:bio", val, s.group(2)

    # tags arrays: tags: ['x','y']
    for m in re.finditer(r"\btags\s*:\s*\[(.*?)\]", raw, re.S):
        for s in re.finditer(r"(['\"])((?:\\\1|(?!\1).)*)\1", m.group(1)):
            val = clean_text(strip_inline_tags(s.group(2)))
            if looks_like_text(val):
                yield section_for(m.start()), "data:tag", val, s.group(2)

    # examples arrays
    for m in re.finditer(r"\bexamples\s*:\s*\[(.*?)\]", raw, re.S):
        for s in re.finditer(r"(['\"])((?:\\\1|(?!\1).)*)\1", m.group(1)):
            val = clean_text(strip_inline_tags(s.group(2)))
            if looks_like_text(val):
                yield section_for(m.start()), "data:example", val, s.group(2)


def collect_for_page(page_cfg):
    rows = []
    seen = set()
    for rel in page_cfg["files"]:
        path = SRC / rel
        if not path.exists():
            continue
        gen = extract_js_data(path) if path.suffix == ".js" else extract_svelte(path)
        for section, etype, text, anchor in gen:
            # Dedup by (text, etype) within page (keep first occurrence)
            key = (section, etype, text)
            if key in seen: continue
            seen.add(key)
            rows.append({
                "section": section,
                "element": etype,
                "text": text,
                "anchor": anchor,
                "file": str(path.relative_to(ROOT)),
            })
    return rows


def write_excel(all_rows, manifest):
    wb = Workbook()
    wb.remove(wb.active)

    HEADER = ["ID", "Section / Block", "Element", "Current Text", "New Text", "Notes"]
    header_fill = PatternFill(start_color="1A8C8E", end_color="1A8C8E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="EFF7F7", end_color="EFF7F7", fill_type="solid")
    section_font = Font(bold=True, color="0C1A2A", size=11)
    thin = Side(border_style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Sheet: Instructions
    ws_inst = wb.create_sheet("README", 0)
    ws_inst.append(["Base One — Content Audit"])
    ws_inst["A1"].font = Font(bold=True, size=18, color="1A8C8E")
    ws_inst.append([])
    ws_inst.append(["How this works:"])
    ws_inst["A3"].font = Font(bold=True, size=12)
    instructions = [
        "1. Each sheet is one page of the website.",
        "2. Rows are grouped by visible section/block (highlighted teal headings).",
        "3. To suggest a change, type the new wording in the 'New Text' column.",
        "4. Leave 'New Text' blank if the current text is fine.",
        "5. Use the 'Notes' column for questions or comments (e.g. \"shorten this\", \"is this fact correct?\").",
        "6. DO NOT change the ID column — that is how we map your edits back to the website.",
        "7. When done, save and send the file back. Modifications will be applied exactly.",
        "",
        "Element column legend:",
        "   <h1>/<h2>/<h3>  →  page title / section heading / subheading",
        "   <p>              →  body paragraph",
        "   <li>             →  list item",
        "   <a>              →  link text",
        "   <button>         →  button label",
        "   <span>           →  inline label or eyebrow",
        "   <blockquote>     →  pulled quote",
        "   prop:heading     →  passed to a reusable component (hero / CTA / etc.)",
        "   attr:alt         →  image alt-text (accessibility / SEO)",
        "   attr:title       →  hover/title attribute",
        "   data:name        →  structured data (team member name, hotel name, etc.)",
        "   data:desc        →  structured data description",
    ]
    for line in instructions:
        ws_inst.append([line])
    ws_inst.column_dimensions["A"].width = 110

    # One sheet per page
    for page in PAGES:
        rows = all_rows[page["key"]]
        ws = wb.create_sheet(page["name"][:31])
        # Header row
        ws.append(HEADER)
        for col, _ in enumerate(HEADER, start=1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(vertical="center", horizontal="left")
            c.border = border
        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A2"

        # Group rows by section, write a section divider row, then data rows
        current_section = None
        idx = 0
        for r in rows:
            if r["section"] != current_section:
                current_section = r["section"]
                ws.append(["", current_section, "", "", "", ""])
                row = ws.max_row
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = section_fill
                    cell.font = section_font
                    cell.border = border
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            idx += 1
            id_ = f"{page['key']}-{idx:03d}"
            ws.append([id_, r["section"], r["element"], r["text"], "", ""])
            row = ws.max_row
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.alignment = wrap
                cell.border = border
            # Light grey ID column
            ws.cell(row=row, column=1).font = Font(name="Menlo", size=10, color="6B7280")

            # Save manifest entry
            manifest[id_] = {
                "page": page["name"],
                "section": r["section"],
                "element": r["element"],
                "file": r["file"],
                "text": r["text"],
                "anchor": r["anchor"],
            }

        # Column widths
        widths = [12, 22, 18, 60, 60, 30]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_XLSX)


def main():
    all_rows = {}
    manifest = {}
    for page in PAGES:
        rows = collect_for_page(page)
        all_rows[page["key"]] = rows
        print(f"  {page['key']:6s} {page['name']:35s} → {len(rows):3d} strings")

    write_excel(all_rows, manifest)
    OUT_MANIFEST.write_text(json.dumps(manifest, indent=2, ensure_ascii=False))
    print(f"\n  ✓ {OUT_XLSX.relative_to(ROOT)}")
    print(f"  ✓ {OUT_MANIFEST.relative_to(ROOT)}  ({len(manifest)} entries)")


if __name__ == "__main__":
    main()

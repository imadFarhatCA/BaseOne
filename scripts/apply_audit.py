#!/usr/bin/env python3
"""Apply auditor edits from content-audit.xlsx back to the source files.

Usage:  python3 apply_audit.py [path/to/edited.xlsx]

Reads the manifest to find each edit's source file + exact anchor, replaces
the original text with the new text, and prints a summary. Modifications
that can't be matched 1:1 are reported so a human can review them.
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("/Users/imadfarhat/imadFarhat projects/Base One")
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "content-audit.xlsx"
MANIFEST = ROOT / "content-audit-manifest.json"


def main():
    manifest = json.load(open(MANIFEST))
    wb = load_workbook(XLSX)

    changes = []  # list of (id, file, old_anchor, new_anchor, old_text, new_text)
    for sheet in wb.sheetnames:
        if sheet == "README": continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            id_, sec, elem, current, new, notes = (list(row) + [None]*6)[:6]
            if not id_ or not new: continue
            new = str(new).strip()
            if not new: continue
            entry = manifest.get(id_)
            if not entry:
                print(f"⚠ Unknown ID: {id_}")
                continue
            old_text = entry["text"]
            if new == old_text: continue   # no change
            # Build new anchor: take old anchor, swap the plain-text portion with new text
            # If the anchor has inline tags (e.g. <br>), try to preserve them when sensible;
            # otherwise just substitute literally.
            old_anchor = entry["anchor"]
            new_anchor = build_new_anchor(old_anchor, old_text, new)
            changes.append({
                "id": id_, "file": entry["file"], "section": entry["section"],
                "element": entry["element"], "old_anchor": old_anchor,
                "new_anchor": new_anchor, "old_text": old_text, "new_text": new,
            })

    if not changes:
        print("No changes to apply.")
        return

    # Group by file and apply
    by_file = {}
    for c in changes:
        by_file.setdefault(c["file"], []).append(c)

    applied, unmatched = 0, []
    for file_rel, edits in by_file.items():
        path = ROOT / file_rel
        text = path.read_text(encoding="utf-8")
        is_js = file_rel.endswith((".js", ".ts"))
        for e in edits:
            if e["old_anchor"] not in text:
                unmatched.append(e)
                print(f"  ⚠ {e['id']:10s} anchor not found verbatim — needs manual edit")
                print(f"     anchor: {e['old_anchor'][:100]}")
                continue
            new_anchor = e["new_anchor"]
            # JS data files: if new text has an apostrophe and the surrounding
            # string literal uses single quotes, rewrite the whole literal as a
            # template literal (backticks) to keep syntax valid.
            if is_js and "'" in new_anchor:
                idx = text.find(e["old_anchor"])
                j = idx - 1
                while j >= 0 and text[j] not in ("'", '"', "`"):
                    j -= 1
                open_q = text[j] if j >= 0 else "'"
                if open_q == "'":
                    k = idx + len(e["old_anchor"])
                    while k < len(text) and text[k] != "'":
                        k += 1
                    new_inner = new_anchor.replace("`", "\\`").replace("${", "\\${")
                    text = text[:j] + "`" + new_inner + "`" + text[k+1:]
                    applied += 1
                    print(f"  ✓ {e['id']:10s} {e['file']:40s}  {e['old_text'][:35]} → {e['new_text'][:35]}  [backtick-wrapped]")
                    continue
            text = text.replace(e["old_anchor"], new_anchor, 1)
            applied += 1
            print(f"  ✓ {e['id']:10s} {e['file']:40s}  {e['old_text'][:35]} → {e['new_text'][:35]}")
        path.write_text(text, encoding="utf-8")

    print(f"\nApplied {applied} change(s). Unmatched: {len(unmatched)}.")


def build_new_anchor(old_anchor: str, old_text: str, new_text: str) -> str:
    """Construct the replacement string preserving inline HTML when possible."""
    # If anchor equals old_text exactly, just use new_text.
    if old_anchor.strip() == old_text.strip():
        return new_text
    # If anchor has <br> or other inline tags, attempt to preserve them by
    # replacing only the textual part. Strategy: tokenize the anchor by tags,
    # then redistribute the new text proportionally? Too risky. Safer: do a
    # plain swap and let the auditor re-add <br> via the New Text column if
    # they want a manual line break. Returning new_text drops embedded tags.
    return new_text


if __name__ == "__main__":
    main()
